import sys
import threading
from tkinter import Tk, filedialog
import io
from PIL import Image
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget, \
    QInputDialog, QMessageBox, QRadioButton, QComboBox, QHBoxLayout, QStyleFactory
from PyQt5.QtCore import Qt, QCoreApplication
from matplotlib.backend_bases import MouseButton
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.patches as patches
import openpyxl
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from qt_material import apply_stylesheet

uri = "mongodb+srv://vovabalaxoncev:Thcvovan7777@cluster0.u499jdc.mongodb.net/?retryWrites=true&w=majority"


class MongoConnection:
    def __init__(self, uri):
        self.client = MongoClient(uri, server_api=ServerApi('1'))
        self.db = self.client['BazaProv']
        self.collection_density = self.db['Rocks']
        self.collection_krepi = self.db['Krepi']
        self.docs = self.collection_density.distinct('Породы')
        self.plot = self.collection_density.distinct('Плотность')
        self.name_krep = self.collection_krepi.distinct('Название')

    def get_docs(self):
        return self.docs

    def get_plot(self):
        return self.plot

    def get_name_krep(self):
        return self.name_krep

    def get_density_collection(self):
        return self.collection_density

    def get_krepi_collection(self):
        return self.collection_krepi


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()

        got_mongo = MongoConnection(uri)
        docs = got_mongo.get_docs()
        name_krep = got_mongo.get_name_krep()
        self.collection_krepi = got_mongo.get_krepi_collection()
        self.collection_density = got_mongo.get_density_collection()

        self.setWindowTitle("Подготовка к расчетам")
        self.setGeometry(100, 100, 800, 800)

        self.label_thickness = QLabel("Толщина:", self)
        self.label_thickness.setGeometry(10, 11, 100, 30)
        self.line_edit_thickness = QLineEdit(self)
        self.line_edit_thickness.setGeometry(30, 10, 30, 30)

        self.label_density = QLabel("Плотность:", self)
        self.label_density.setGeometry(10, 130, 100, 30)

        self.label_part = QLabel("Длина участка в метрах:", self)
        self.label_part.setGeometry(10, 90, 150, 30)
        self.line_edit_part = QLineEdit(self)
        self.line_edit_part.setGeometry(30, 90, 30, 30)

        self.adder = Adder
        self.button_add = QPushButton("Добавить слой", self)
        self.button_add.setGeometry(10, 210, 210, 30)
        self.button_add.clicked.connect(lambda: self.adder.add_layer(self))

        self.button_clear = QPushButton("Сброс", self)
        self.button_clear.setGeometry(10, 210, 210, 30)
        self.button_clear.clicked.connect(self.clear_all)
        self.button_clear.setProperty('class', 'danger')

        self.lines = Lines
        self.button_auto_part = QPushButton("Автоматическое разбиение", self)
        self.button_auto_part.setGeometry(10, 250, 210, 30)
        self.button_auto_part.clicked.connect(lambda: self.lines.auto_part(self))

        self.button_undo = QPushButton("Отмена", self)
        self.button_undo.setGeometry(10, 290, 210, 30)
        self.button_undo.clicked.connect(self.undo_action)

        self.grid = Grid
        self.button_edit_grid = QPushButton("Подготовка к редактированию и сохранению", self)
        self.button_edit_grid.setGeometry(10, 210, 210, 30)
        self.button_edit_grid.clicked.connect(lambda: self.grid.create_grid(self))

        save = Save()
        self.info_layers = []
        self.triangle_coordinates = []
        self.button_save_1 = QPushButton("Сохранить", self)
        self.button_save_1.setGeometry(10, 330, 210, 30)
        self.button_save_1.clicked.connect(
            lambda: save.save_grid_information(self.triangle_coordinates, self.info_layers))

        self.radio_vertical = QRadioButton("Вертикальная линия", self)
        self.radio_vertical.setGeometry(10, 370, 150, 40)
        self.radio_vertical.toggled.connect(self.toggle_vertical_line)

        self.radio_horizontal = QRadioButton("Горизонтальная линия", self)
        self.radio_horizontal.setGeometry(10, 410, 150, 40)
        self.radio_horizontal.toggled.connect(self.toggle_horizontal_line)

        self.radio_edit = QRadioButton("Режим редактирования сетки", self)
        self.radio_edit.setGeometry(10, 410, 150, 40)
        self.radio_edit.toggled.connect(self.toggle_edit)

        # слои
        self.combobox_layer = QComboBox()
        self.combobox_layer.addItems(docs)
        self.combobox_layer.setStyleSheet("color: white;")

        self.label_krep = QLabel("Крепь:", self)
        self.combobox_krep = QComboBox()
        self.combobox_krep.addItems(name_krep)
        self.combobox_krep.setStyleSheet("color: white;")

        self.button_add_krep = QPushButton("Добавить крепи автоматически", self)
        self.button_add_krep.setGeometry(10, 330, 210, 30)
        self.button_add_krep.clicked.connect(lambda: self.adder.add_krep(self, None))

        self.radio_krep = QRadioButton("Добавить крепи", self)
        self.radio_krep.setGeometry(10, 370, 150, 40)
        self.radio_krep.toggled.connect(self.toggle_krep)

        self.label_width_krep = QLabel("Расстояние между крепями:", self)
        self.label_width_krep.setGeometry(10, 90, 150, 30)
        self.line_edit_width_krep = QLineEdit(self)
        self.line_edit_width_krep.setGeometry(30, 90, 30, 30)

        self.label_coordinate = QLabel("Координаты:", self)

        # объект для графика
        self.figure = Figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        self.axes = self.figure.add_subplot(111)
        self.figure.set_animated(True)
        self.background = self.canvas.copy_from_bbox(self.axes.bbox)

        # список для хранения слоев
        self.rectangles = []

        # список для хранения разбиений
        self.partitions = []

        # компоновщики
        layout = QVBoxLayout()
        krepi_layout = QHBoxLayout()
        top_panel_layout = QHBoxLayout()
        graph_panel_layout = QHBoxLayout()

        top_panel_layout.addWidget(self.combobox_layer)
        top_panel_layout.addWidget(self.label_thickness)
        top_panel_layout.addWidget(self.line_edit_thickness)
        top_panel_layout.addWidget(self.label_part)
        top_panel_layout.addWidget(self.line_edit_part)
        top_panel_layout.addWidget(self.label_density)
        graph_panel_layout.addWidget(self.label_coordinate)
        graph_panel_layout.addWidget(self.radio_vertical)
        graph_panel_layout.addWidget(self.radio_horizontal)
        graph_panel_layout.addWidget(self.radio_edit)
        top_panel_layout.addWidget(self.button_add)
        graph_panel_layout.addWidget(self.button_auto_part)
        top_panel_layout.addWidget(self.button_undo)
        graph_panel_layout.addWidget(self.button_edit_grid)
        top_panel_layout.addWidget(self.button_save_1)
        top_panel_layout.addWidget(self.button_clear)
        krepi_layout.addWidget(self.label_krep)
        krepi_layout.addWidget(self.combobox_krep)
        krepi_layout.addWidget(self.radio_krep)
        krepi_layout.addWidget(self.label_width_krep)
        krepi_layout.addWidget(self.line_edit_width_krep)
        krepi_layout.addWidget(self.button_add_krep)
        layout.addLayout(top_panel_layout)
        layout.addLayout(krepi_layout)
        layout.addLayout(graph_panel_layout)
        layout.addWidget(self.canvas)

        # виджет для размещения компоновки
        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)
        self.final_Y = 0
        self.check_first = 0

        self.is_width_locked = False
        self.auto_save_enabled = False
        self.net_enabled = False

        # обработчики событий мыши на FigureCanvas
        self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self.canvas.mpl_connect("motion_notify_event", self.on_motion_notify)
        self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.canvas.mpl_connect('scroll_event', self.on_scroll)

        self.first_coordinate_line_y = 0
        self.horizontal_lines = []  # Список координат горизонтальных линий
        self.vertical_lines = []  # Список координат вертикальных линий
        self.info_krep = []
        self.first_line_horizontal_check = True
        self.first_line_vertical_check = True
        self.dragging = False
        self.x_grid_edit = 0
        self.y_grid_edit = 0
        self.current_x = 0
        self.current_y = 0
        self.pressed = False
        self.prev_x = None
        self.prev_y = None
        self.translation_factor = 0.005  # Коэффициент перемещения
        self.current_x_lim = None
        self.current_y_lim = None
        self.current_partitions = []
        self.krep_coordinate = []
        self.index_edit_triangle = []

        self.draw_graph = DrawFigure

    def on_mouse_press(self, event):
        if event.button == MouseButton.MIDDLE:
            self.pressed = True
            self.prev_x = event.x
            self.prev_y = event.y
        if self.radio_edit.isChecked() and event.button == MouseButton.LEFT:
            self.grid.grid_edit(self, event.xdata, event.ydata)
        if self.radio_krep.isChecked() and self.rectangles and event.button == MouseButton.LEFT:
            self.adder.add_krep(self, event.xdata)
        if event.button == MouseButton.LEFT:
            if self.info_layers:
                self.lines.add_line(self, event.xdata, event.ydata)
            else:
                msg = QMessageBox()
                msg.setWindowTitle("Ошибка")
                msg.setText("Возникли проблемы со слоями!")
                msg.setIcon(QMessageBox.Warning)
                msg.exec_()

    def on_scroll(self, event):
        if event.button == 'up':
            # Увеличение масштаба при прокрутке вверх
            self.current_x_lim = [self.axes.get_xlim()[0] * 0.9, self.axes.get_xlim()[1] * 0.9]
            self.current_y_lim = [self.axes.get_ylim()[0] * 0.9, self.axes.get_ylim()[1] * 0.9]
            self.axes.set_xlim(self.axes.get_xlim()[0] * 0.9, self.axes.get_xlim()[1] * 0.9)
            self.axes.set_ylim(self.axes.get_ylim()[0] * 0.9, self.axes.get_ylim()[1] * 0.9)
        elif event.button == 'down':
            # Уменьшение масштаба при прокрутке вниз
            self.current_x_lim = [self.axes.get_xlim()[0] * 1.1, self.axes.get_xlim()[1] * 1.1]
            self.current_y_lim = [self.axes.get_ylim()[0] * 1.1, self.axes.get_ylim()[1] * 1.1]
            self.axes.set_xlim(self.axes.get_xlim()[0] * 1.1, self.axes.get_xlim()[1] * 1.1)
            self.axes.set_ylim(self.axes.get_ylim()[0] * 1.1, self.axes.get_ylim()[1] * 1.1)

        self.canvas.blit(self.axes.bbox)
        self.canvas.draw()

    def on_motion_notify(self, event):
        if event.xdata is not None:
            self.label_coordinate.setText(f'Координаты: x {round(event.xdata, 2)} y {round(event.ydata, 2)}')
        if self.pressed:
            translation_factor_x = self.axes.get_xlim()[1] / 1000
            translation_factor_y = self.axes.get_ylim()[1] / 1000
            dx = (event.x - self.prev_x) * translation_factor_x
            dy = (event.y - self.prev_y) * translation_factor_y
            self.current_x_lim = [self.axes.get_xlim()[0] - dx, self.axes.get_xlim()[1] - dx]
            self.current_y_lim = [self.axes.get_ylim()[0] + dy, self.axes.get_ylim()[1] + dy]
            self.axes.set_xlim(self.axes.get_xlim()[0] - dx, self.axes.get_xlim()[1] - dx)
            self.axes.set_ylim(self.axes.get_ylim()[0] + dy, self.axes.get_ylim()[1] + dy)
            self.prev_x = event.x
            self.prev_y = event.y
            self.canvas.draw()
        if self.radio_edit.isChecked() and self.dragging:

            x = event.xdata
            y = event.ydata
            x_grid_edit = self.x_grid_edit
            y_grid_edit = self.y_grid_edit
            if float(self.line_edit_part.text()) > 500:
                radius = 0.1
                y_grid_edit = round(y_grid_edit, 1)
                x_grid_edit = round(x_grid_edit, 1)

            else:
                radius = 0.1
                y_grid_edit = round(y_grid_edit, 1)
                x_grid_edit = round(x_grid_edit, 1)
            for index in self.index_edit_triangle:
                triangle = self.triangle_coordinates[index]
                if float(self.line_edit_part.text()) > 500:
                    x0 = round(triangle['x0'], 1)
                    x1 = round(triangle['x1'], 1)
                    x2 = round(triangle['x2'], 1)
                    y0 = round(triangle['y0'], 1)
                    y1 = round(triangle['y1'], 1)
                    y2 = round(triangle['y2'], 1)
                else:
                    x0 = round(triangle['x0'], 1)
                    x1 = round(triangle['x1'], 1)
                    x2 = round(triangle['x2'], 1)
                    y0 = round(triangle['y0'], 1)
                    y1 = round(triangle['y1'], 1)
                    y2 = round(triangle['y2'], 1)
                if (y0 == y_grid_edit and x0 == x_grid_edit) or (
                        y0 == y_grid_edit + radius and x0 == x_grid_edit + radius) or (
                        y0 == y_grid_edit - radius and x0 == x_grid_edit - radius) or (
                        y0 == y_grid_edit - radius and x0 == x_grid_edit + radius) or (
                        y0 == y_grid_edit + radius and x0 == x_grid_edit - radius):
                    triangle['y0'] = y
                    triangle['x0'] = x
                    self.y_grid_edit = y
                    self.x_grid_edit = x

                if (y1 == y_grid_edit and x1 == x_grid_edit) or (
                        y1 == y_grid_edit + radius and x1 == x_grid_edit + radius) or (
                        y1 == y_grid_edit - radius and x1 == x_grid_edit - radius) or (
                        y1 == y_grid_edit - radius and x1 == x_grid_edit + radius) or (
                        y1 == y_grid_edit + radius and x1 == x_grid_edit - radius):
                    triangle['y1'] = y
                    triangle['x1'] = x
                    self.y_grid_edit = y
                    self.x_grid_edit = x

                if (y2 == y_grid_edit and x2 == x_grid_edit) or (
                        y2 == y_grid_edit + radius and x2 == x_grid_edit + radius) or (
                        y2 == y_grid_edit - radius and x2 == x_grid_edit - radius) or (
                        y2 == y_grid_edit - radius and x2 == x_grid_edit + radius) or (
                        y2 == y_grid_edit + radius and x2 == x_grid_edit - radius):
                    triangle['y2'] = y
                    triangle['x2'] = x
                    self.y_grid_edit = y
                    self.x_grid_edit = x

            for index in self.index_edit_triangle:
                triangle = self.triangle_coordinates[index]
                x = [triangle['x0'], triangle['x1'], triangle['x2']]
                y = [triangle['y0'], triangle['y1'], triangle['y2']]
                line = self.axes.lines[index]
                line.set_data(x + [x[0]], y + [y[0]])

                # Перерисовать только измененную линию или точку
                self.axes.draw_artist(line)
            self.canvas.blit(self.axes.bbox)
            self.canvas.draw()

    def on_mouse_release(self, event):
        if event.button == MouseButton.MIDDLE:
            self.pressed = False
        self.dragging = False

    def toggle_vertical_line(self, checked):
        return

    def toggle_krep(self, checked):
        return

    def toggle_horizontal_line(self, checked):
        return

    def toggle_edit(self, checked):
        return

    def undo_action(self):
        if self.rectangles:
            self.info_krep = []
            self.rectangles.pop()
            self.draw_graph.draw_layer(self)
            self.final_Y = self.final_Y - float(self.line_edit_thickness.text())
            self.current_partitions.pop()
            self.horizontal_lines = []
            self.vertical_lines = []
            self.triangle_coordinates = []
            self.first_line_horizontal_check = True
            self.first_line_vertical_check = True

    def clear_all(self):
        self.is_width_locked = False
        self.auto_save_enabled = False
        self.net_enabled = False

        self.first_coordinate_line_y = 0
        self.horizontal_lines = []
        self.vertical_lines = []
        self.final_Y = 0
        self.check_first = 0
        self.info_krep = []
        self.current_partitions = []
        self.first_line_horizontal_check = True
        self.first_line_vertical_check = True

        self.dragging = False
        self.x_grid_edit = 0
        self.y_grid_edit = 0
        self.current_x = 0
        self.current_y = 0

        self.pressed = False
        self.prev_x = None
        self.prev_y = None
        self.current_x_lim = None
        self.current_y_lim = None

        self.krep_coordinate = []
        self.index_edit_triangle = []
        self.rectangles = []

        # список для хранения разбиений
        self.partitions = []
        self.info_layers = []
        self.triangle_coordinates = []
        self.axes.clear()

        self.line_edit_thickness.setText('')
        self.line_edit_width_krep.setText('')
        self.line_edit_part.setText('')
        self.line_edit_part.setReadOnly(False)
        self.canvas.draw()


class Adder:
    def add_krep(self, xdata):
        self.line_edit_part.setReadOnly(True)
        if self.info_layers:
            if self.radio_krep.isChecked() and xdata:
                all_width = xdata + 1
                distance = 0
                coordinate_x1_distance = xdata
                coordinate_final = 0
            if (self.line_edit_part.text()).isdigit() and (self.line_edit_width_krep.text()).isdigit():
                all_width = float(self.line_edit_part.text())
                distance = float(self.line_edit_width_krep.text())
                coordinate_x1_distance = 0
                coordinate_final = 0
            if not ((self.line_edit_part.text()).isdigit() or (self.line_edit_width_krep.text()).isdigit()) and not (
            self.radio_krep.isChecked()):
                msg = QMessageBox()
                msg.setWindowTitle("Ошибка")
                msg.setText("Возникли проблемы со слоями!")
                msg.setIcon(QMessageBox.Warning)
                msg.exec_()
                return

            result = self.collection_krepi.find({'Название': str(self.combobox_krep.currentText())})
            image_data = None
            long = None
            for doc in result:
                image_data = doc['Изображение']
                long = doc['Козырек длина']
                height = doc['Высота']
                point1_x = (doc['Точка в.л.у'])[0]
                point1_y = (doc['Точка в.л.у'])[1]
                point2_x = (doc['Точка в.л.у.2'])[0]
                point2_y = (doc['Точка в.л.у.2'])[1]
                point3_x = (doc['Точка в.п.у'])[0]
                point3_y = (doc['Точка в.п.у'])[1]
                point4_x = (doc['Точка в.п.у.2'])[0]
                point4_y = (doc['Точка в.п.у.2'])[1]
                point5_x = (doc['Точка н.л.у'])[0]
                point5_y = (doc['Точка н.л.у'])[1]
                point6_x = (doc['Точка н.л.у.2'])[0]
                point6_y = (doc['Точка н.л.у.2'])[1]
                point7_x = (doc['Точка н.п.у'])[0]
                point7_y = (doc['Точка н.п.у'])[1]
                point8_x = (doc['Точка н.п.у.2'])[0]
                point8_y = (doc['Точка н.п.у.2'])[1]

            while coordinate_final < all_width:
                layer = self.info_layers[0]

                img = Image.open(io.BytesIO(image_data))
                img = img.convert('RGBA')
                x1 = coordinate_x1_distance
                x = x1 + long
                y1 = layer['y0']
                y = layer['y1']
                edit_height = abs(abs(y1) - abs(y) - height)
                y1_edit = y1 - edit_height

                point1_x += x1
                point1_y += y1_edit
                point2_x += x1
                point2_y += y1_edit
                point3_x += x1
                point3_y += y1_edit
                point4_x += x1
                point4_y += y1_edit
                point5_x += x1
                point5_y += y1
                point6_x += x1
                point6_y += y1
                point7_x += x1
                point7_y += y1
                point8_x += x1
                point8_y += y1

                coordinate_info = {
                    'название': str(self.combobox_krep.currentText()),
                    'x1': x1,
                    'x': x,
                    'y1': y1,
                    'y': y,
                    'image': img,
                    'point1_x': point1_x,
                    'point1_y': point1_y,
                    'point2_x': point2_x,
                    'point2_y': point2_y,
                    'point3_x': point3_x,
                    'point3_y': point3_y,
                    'point4_x': point4_x,
                    'point4_y': point4_y,
                    'point5_x': point5_x,
                    'point5_y': point5_y,
                    'point6_x': point6_x,
                    'point6_y': point6_y,
                    'point7_x': point7_x,
                    'point7_y': point7_y,
                    'point8_x': point8_x,
                    'point8_y': point8_y,
                }
                self.info_krep.append(coordinate_info)
                self.axes.imshow(img, extent=([x1, x, y1, y]), aspect='equal', zorder=10)
                # x_min, x_max = self.current_x_lim
                # y_min, y_max = self.current_y_lim
                # self.axes.set_xlim([x_min, x_max])
                # self.axes.set_ylim(
                #     [y_min + (x_max / 2), y_max + (x_max / 2)])
                coordinate_x1_distance += distance + long
                coordinate_final = coordinate_x1_distance + distance + long
            self.figure.tight_layout()
            self.axes.set_aspect('auto')
            self.canvas.draw()
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("Возникли проблемы со слоями!")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()
            return

    def add_layer(self):
        rect_name = self.combobox_layer.currentText()

        self.info_krep = []
        self.horizontal_lines = []
        self.vertical_lines = []
        self.triangle_coordinates = []
        self.first_line_horizontal_check = True
        self.first_line_vertical_check = True
        self.current_x_lim = [self.axes.get_xlim()[0] * 0.9, self.axes.get_xlim()[1] * 0.9]
        self.current_y_lim = [self.axes.get_ylim()[0] * 0.9, self.axes.get_ylim()[1] * 0.9]
        if (self.line_edit_thickness.text()).isdigit():
            height = float(self.line_edit_thickness.text())
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("Некорректное значение")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()
            return
        if (self.line_edit_part.text()).isdigit():
            width = float(self.line_edit_part.text())
            self.line_edit_part.setReadOnly(True)
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("Некорректное значение")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()
            return
        result = self.collection_density.find({'Породы': str(self.combobox_layer.currentText())})
        rect_color = None
        density = None
        for doc in result:
            density = doc['Плотность']
            rect_color = doc['Цвет']
        self.label_density.setText(f'Плотность: {density}')

        if self.check_first != 0:
            x = 0
            y = self.final_Y
            self.final_Y = y + height
        if self.check_first == 0:
            x = 0
            y = 0 - height
            self.check_first += 1
            self.first_coordinate_line_y = y

        rect = patches.Rectangle((x, y), width, height, facecolor=rect_color)
        rect.set_label(rect_name)
        self.rectangles.append(rect)
        self.draw_graph.draw_layer(self)


        # Добавление информации о слое в текущее разбиение
        current_partition = {
            'name': rect_name,
            'density': density,
            'thickness': height,
            'color': rect_color
        }
        self.current_partitions.append(current_partition)
        layer_info = {
            'name': rect_name,
            'x0': x,
            'y0': y,
            'x1': x + width,
            'y1': y + height,
            'color': rect_color,
            'density': density
        }
        self.info_layers.append(layer_info)


class DrawFigure:
    def draw_layer(self):
        if self.radio_edit.isChecked():
            self.axes.clear()
            for index, info_krep in enumerate(self.info_krep):
                self.axes.imshow(info_krep['image'],
                                 extent=([info_krep['x1'], info_krep['x'], info_krep['y1'], info_krep['y']]),
                                 aspect='equal', zorder=10)
            for rect in self.rectangles:
                self.axes.add_patch(rect)
                self.axes.annotate(rect.get_label(),
                                   (rect.get_x() + rect.get_width() / 2, rect.get_y() + rect.get_height() / 2),
                                   ha='center', va='center')
            x_min, x_max = self.current_x_lim
            y_min, y_max = self.current_y_lim
            self.axes.set_xlim([x_min, x_max])
            self.axes.set_ylim(
                [y_min, y_max])
            self.figure.tight_layout()
            self.axes.set_aspect('auto')
            self.canvas.draw()
        if not self.radio_edit.isChecked():
            self.axes.clear()
            for rect in self.rectangles:
                self.axes.add_patch(rect)
                self.axes.annotate(rect.get_label(),
                                   (rect.get_x() + rect.get_width() / 2, rect.get_y() + rect.get_height() / 2),
                                   ha='center', va='center')
            self.current_x_lim = [0, float(self.line_edit_part.text())]
            self.current_y_lim = [0 - float(self.line_edit_thickness.text()) * 2,
                                  self.final_Y + float(self.line_edit_thickness.text())]
            self.axes.set_xlim([0, float(self.line_edit_part.text())])
            self.axes.set_ylim(
                [0 - float(self.line_edit_thickness.text()) * 2,
                 self.final_Y + float(self.line_edit_thickness.text())])
            self.figure.tight_layout()
            self.axes.set_aspect('auto')
            self.canvas.draw()

    def draw_triangles(self, all_coordinates):
        number_triangle = 0
        for coordinates in all_coordinates:
            number_triangle += 1
            triangle_info = {
                'number': number_triangle,
                'x0': coordinates[1],
                'x1': coordinates[3],
                'x2': coordinates[1],
                'y0': coordinates[0],
                'y1': coordinates[2],
                'y2': coordinates[2],

            }
            number_triangle += 1
            self.triangle_coordinates.append(triangle_info)
            triangle_info = {
                'number': number_triangle,
                'x0': coordinates[1],
                'x1': coordinates[3],
                'x2': coordinates[3],
                'y0': coordinates[0],
                'y1': coordinates[2],
                'y2': coordinates[0],

            }
            self.triangle_coordinates.append(triangle_info)
        draw = DrawFigure
        draw.draw_layer(self)
        # self.canvas.draw()
        info_krep_images = []
        for index, info_krep in enumerate(self.info_krep):
            info_krep_images.append(info_krep['image'])
            self.axes.imshow(info_krep['image'],
                             extent=([info_krep['x1'], info_krep['x'], info_krep['y1'], info_krep['y']]),
                             aspect='equal', zorder=10)

        for index, triangle in enumerate(self.triangle_coordinates):
            x_coords = [triangle['x0'], triangle['x1'], triangle['x2']]
            y_coords = [triangle['y0'], triangle['y1'], triangle['y2']]
            self.axes.plot(x_coords + [x_coords[0]], y_coords + [y_coords[0]], color='red', zorder=11)

        self.figure.tight_layout()
        self.axes.set_aspect('auto')
        self.canvas.draw()


class Grid:
    def create_grid(self):
        if self.vertical_lines and self.horizontal_lines:
            self.vertical_lines = sorted(self.vertical_lines, key=lambda line: line['x0'])
            self.horizontal_lines = sorted(self.horizontal_lines, key=lambda line: line['y0'])

            vert_wall = []
            horiz_wall = []

            first_w = 0
            x_pred = 0

            for line_vertical in self.vertical_lines:

                x1_vert = line_vertical['x1']

                if x1_vert <= x_pred:
                    x_pred = x1_vert
                    first_w += 1
                    continue
                if first_w > 0:
                    vert_wall.append([x_pred, x1_vert])
                    x_pred = x1_vert
                    continue
                first_w += 1

            first_w = 0
            y_pred = 0
            for line_horizontal in self.horizontal_lines:

                y1_hor = line_horizontal['y1']

                if y1_hor <= y_pred:
                    y_pred = y1_hor
                    first_w += 1
                    continue
                if first_w > 0:
                    horiz_wall.append([y_pred, y1_hor])
                    y_pred = y1_hor
                    continue
                first_w += 1

            all_coordinates = []
            for line_horizontal in horiz_wall:
                for line_vertical in vert_wall:
                    all_coordinates.append([line_horizontal[0], line_vertical[0], line_horizontal[1], line_vertical[1]])
            self.draw_graph.draw_triangles(self, all_coordinates)

        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("Возникли проблемы со линиями сетки!")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()

    def grid_edit(self, xdata, ydata):
        self.index_edit_triangle = []
        if float(self.line_edit_part.text()) > 500:
            radius = 0.1
            x = round(xdata, 1)
            y = round(ydata, 1)
        else:
            radius = 0.1
            x = round(xdata, 1)
            y = round(ydata, 1)
        for index, coordinate in enumerate(self.triangle_coordinates):
            if float(self.line_edit_part.text()) > 500:
                x0 = round(coordinate['x0'], 1)
                x1 = round(coordinate['x1'], 1)
                x2 = round(coordinate['x2'], 1)
                y0 = round(coordinate['y0'], 1)
                y1 = round(coordinate['y1'], 1)
                y2 = round(coordinate['y2'], 1)
            else:
                x0 = round(coordinate['x0'], 1)
                x1 = round(coordinate['x1'], 1)
                x2 = round(coordinate['x2'], 1)
                y0 = round(coordinate['y0'], 1)
                y1 = round(coordinate['y1'], 1)
                y2 = round(coordinate['y2'], 1)
            if (
                    x == x0 or x == x1 or x == x2 or x == x0 + radius or x == x1 + radius or x == x2 + radius or x == x0 - radius or x == x1 - radius or x == x2 - radius) and (
                    y == y0 or y == y1 or y == y2 or y == y0 + radius or y == y1 + radius or y == y2 + radius or y == y0 - radius or y == y1 - radius or y == y2 - radius):
                self.dragging = True
                self.x_grid_edit = x
                self.y_grid_edit = y
                self.index_edit_triangle.append(index)


class Lines:
    def add_line(self, xdata, ydata):
        x = xdata
        y = ydata
        if x is not None and y is not None:
            if (self.line_edit_part.text()).isdigit():
                height = int(self.line_edit_part.text())
            else:
                msg = QMessageBox()
                msg.setWindowTitle("Ошибка")
                msg.setText("Для начала постройте слои!")
                msg.setIcon(QMessageBox.Warning)
                msg.exec_()
                return
            height_sum = sum(partition['thickness'] for partition in self.current_partitions)
            ylim = height_sum
            if self.radio_horizontal.isChecked():
                if self.first_line_horizontal_check:
                    if self.info_krep:
                        info_krep = self.info_krep[0]
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': info_krep['point8_y'],
                            'y1': info_krep['point8_y']
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height], [info_krep['point8_y'],
                                                     info_krep['point8_y']], color='red')
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': info_krep['point1_y'],
                            'y1': info_krep['point1_y']
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height], [info_krep['point1_y'],
                                                     info_krep['point1_y']], color='red')

                    line_info = {
                        'x0': 0,
                        'x1': height,
                        'y0': self.first_coordinate_line_y + ylim,
                        'y1': self.first_coordinate_line_y + ylim
                    }
                    self.horizontal_lines.append(line_info)

                    self.axes.plot([0, height], [self.first_coordinate_line_y + ylim,
                                                 self.first_coordinate_line_y + ylim], color='red')
                    for layer in self.info_layers:
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': layer['y0'],
                            'y1': layer['y0']
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height], [layer['y0'], layer['y0']], color='red')
                    self.first_line_horizontal_check = False

                line_info = {
                    'x0': 0,
                    'x1': height,
                    'y0': y,
                    'y1': y
                }
                self.horizontal_lines.append(line_info)

                self.axes.plot([0, height], [y, y], color='red')
                self.canvas.draw()

            if self.radio_vertical.isChecked():
                if self.first_line_vertical_check:
                    if self.info_krep:
                        for index, info_krep in enumerate(self.info_krep):
                            for i in range(4):
                                key = 'point{}_x'.format(i + 1)
                                line_info = {
                                    'x0': info_krep[key],
                                    'x1': info_krep[key],
                                    'y0': self.first_coordinate_line_y,
                                    'y1': self.first_coordinate_line_y + ylim
                                }
                                self.vertical_lines.append(line_info)

                                self.axes.plot([info_krep[key], info_krep[key]],
                                               [self.first_coordinate_line_y,
                                                self.first_coordinate_line_y + ylim],
                                               color='blue')

                    line_info = {
                        'x0': 0,
                        'x1': 0,
                        'y0': self.first_coordinate_line_y,
                        'y1': self.first_coordinate_line_y + ylim
                    }
                    self.vertical_lines.append(line_info)

                    self.axes.plot([0, 0],
                                   [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                   color='blue')
                    line_info = {
                        'x0': height,
                        'x1': height,
                        'y0': self.first_coordinate_line_y,
                        'y1': self.first_coordinate_line_y + ylim
                    }
                    self.vertical_lines.append(line_info)

                    self.axes.plot([height, height],
                                   [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                   color='blue')
                    self.first_line_vertical_check = False

                line_info = {
                    'x0': x,
                    'x1': x,
                    'y0': self.first_coordinate_line_y,
                    'y1': self.first_coordinate_line_y + ylim
                }
                self.vertical_lines.append(line_info)

                self.axes.plot([x, x], [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                               color='blue')
                self.canvas.draw()
        pass

    def auto_part(self):
        if self.info_layers:
            if (self.line_edit_part.text()).isdigit():
                height = int(self.line_edit_part.text())
            else:
                msg = QMessageBox()
                msg.setWindowTitle("Ошибка")
                msg.setText("Для начала постройте слои!")
                msg.setIcon(QMessageBox.Warning)
                msg.exec_()
                return
            distance, ok = QInputDialog.getDouble(self, "Расстояние разбиения", "Расстояние:")
            if not ok:
                return
            full_height = abs(self.info_layers[0]['y0']) + abs(self.info_layers[len(self.info_layers) - 1]['y1'])
            full_width = abs(self.info_layers[0]['x0']) + abs(self.info_layers[len(self.info_layers) - 1]['x1'])
            if ok:
                height_sum = sum(partition['thickness'] for partition in self.current_partitions)
                ylim = height_sum
                x = 0
                y = self.first_coordinate_line_y
                for i in range(int(ylim)):
                    if self.first_line_horizontal_check:
                        if self.info_krep:
                            info_krep = self.info_krep[0]
                            line_info = {
                                'x0': 0,
                                'x1': height,
                                'y0': info_krep['point8_y'],
                                'y1': info_krep['point8_y']
                            }
                            self.horizontal_lines.append(line_info)

                            self.axes.plot([0, height], [info_krep['point8_y'],
                                                         info_krep['point8_y']], color='red')
                            line_info = {
                                'x0': 0,
                                'x1': height,
                                'y0': info_krep['point1_y'],
                                'y1': info_krep['point1_y']
                            }
                            self.horizontal_lines.append(line_info)

                            self.axes.plot([0, height], [info_krep['point1_y'],
                                                         info_krep['point1_y']], color='red')
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': self.first_coordinate_line_y,
                            'y1': self.first_coordinate_line_y
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height], [self.first_coordinate_line_y, self.first_coordinate_line_y],
                                       color='red')
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': self.first_coordinate_line_y + ylim,
                            'y1': self.first_coordinate_line_y + ylim
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height],
                                       [self.first_coordinate_line_y + ylim, self.first_coordinate_line_y + ylim],
                                       color='red')
                        for layer in self.info_layers:
                            line_info = {
                                'x0': 0,
                                'x1': height,
                                'y0': layer['y0'],
                                'y1': layer['y0']
                            }
                            self.horizontal_lines.append(line_info)

                            self.axes.plot([0, height], [layer['y0'], layer['y0']], color='red')
                        self.first_line_horizontal_check = False
                    if y + distance <= full_height:
                        line_info = {
                            'x0': 0,
                            'x1': height,
                            'y0': y,
                            'y1': y
                        }
                        self.horizontal_lines.append(line_info)

                        self.axes.plot([0, height], [y, y], color='red')
                        y += distance

                for i in range(height):
                    if self.first_line_vertical_check:
                        if self.info_krep:
                            for index, info_krep in enumerate(self.info_krep):
                                for i in range(4):
                                    key = 'point{}_x'.format(i + 1)
                                    line_info = {
                                        'x0': info_krep[key],
                                        'x1': info_krep[key],
                                        'y0': self.first_coordinate_line_y,
                                        'y1': self.first_coordinate_line_y + ylim
                                    }
                                    self.vertical_lines.append(line_info)

                                    self.axes.plot([info_krep[key], info_krep[key]],
                                                   [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                                   color='blue')
                        line_info = {
                            'x0': 0,
                            'x1': 0,
                            'y0': self.first_coordinate_line_y,
                            'y1': self.first_coordinate_line_y + ylim
                        }
                        self.vertical_lines.append(line_info)

                        self.axes.plot([0, 0], [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                       color='blue')
                        line_info = {
                            'x0': height,
                            'x1': height,
                            'y0': self.first_coordinate_line_y,
                            'y1': self.first_coordinate_line_y + ylim
                        }
                        self.vertical_lines.append(line_info)

                        self.axes.plot([height, height],
                                       [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                       color='blue')
                        self.first_line_vertical_check = False
                    if x + distance <= full_width:
                        line_info = {
                            'x0': x,
                            'x1': x,
                            'y0': self.first_coordinate_line_y,
                            'y1': self.first_coordinate_line_y + ylim
                        }
                        self.vertical_lines.append(line_info)

                        self.axes.plot([x, x], [self.first_coordinate_line_y, self.first_coordinate_line_y + ylim],
                                       color='blue')
                        x += distance
                self.canvas.draw()
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("Возникли проблемы со слоями!")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()


class Save:
    def save_grid_information(self, triangle_coordinates, info_layers):
        if triangle_coordinates:
            cells = []
            for layer in info_layers:
                x0_layer = layer['x0']
                x1_layer = layer['x1']
                y0_layer = layer['y0']
                y1_layer = layer['y1']
                for index, coordinates in enumerate(triangle_coordinates):
                    if coordinates['x0'] >= x0_layer and coordinates['y0'] >= y0_layer and coordinates[
                        'x1'] <= x1_layer and coordinates['y1'] <= y1_layer and x0_layer <= coordinates[
                        'x2'] <= x1_layer and y0_layer <= \
                            coordinates['y2'] <= y1_layer:
                        cell_info = {
                            'name': layer['name'],
                            'color': layer['color'],
                            'x0': coordinates['x0'],
                            'x1': coordinates['x1'],
                            'x2': coordinates['x2'],
                            'y0': coordinates['y0'],
                            'y1': coordinates['y1'],
                            'y2': coordinates['y2'],
                            'density': layer['density']

                        }
                        cells.append(cell_info)

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            headers = ['Name', 'Color', 'x0', 'x1', 'x2', 'y0', 'y1', 'y2', 'density']

            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col).value = header

            for index, cell_info in enumerate(cells, start=2):
                sheet.cell(row=index, column=1).value = cell_info['name']
                sheet.cell(row=index, column=2).value = cell_info['color']
                sheet.cell(row=index, column=3).value = cell_info['x0']
                sheet.cell(row=index, column=4).value = cell_info['x1']
                sheet.cell(row=index, column=5).value = cell_info['x2']
                sheet.cell(row=index, column=6).value = cell_info['y0']
                sheet.cell(row=index, column=7).value = cell_info['y1']
                sheet.cell(row=index, column=8).value = cell_info['y2']
                sheet.cell(row=index, column=9).value = cell_info['density']

            root = Tk()
            root.withdraw()
            file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')

            if file_path:

                workbook.save(file_path)
                print("Excel file saved successfully.")
            else:
                print("Save operation canceled.")

            root.destroy()
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Ошибка")
            msg.setText("кажется нечего сохранять!")
            msg.setIcon(QMessageBox.Warning)
            msg.exec_()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    apply_stylesheet(app, theme='dark_blue.xml')
    app.setWindowIcon(QIcon("icon.png"))
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
